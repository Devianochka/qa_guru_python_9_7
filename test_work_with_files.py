'''Задание
– Запаковать в zip архив несколько разных файлов: pdf, xlsx, csv;

– Положить его в ресурсы;

– Реализовать чтение и проверку содержимого каждого файла из архива
'''


import os
from script_os import resources, path, zip_path
import zipfile
from pypdf import PdfReader
from io import TextIOWrapper
from openpyxl import load_workbook
import csv
import pytest


@pytest.fixture
def arc_files():
    if not os.path.exists(resources):
        os.mkdir(resources)
    file_dir = os.listdir(path)
    with zipfile.ZipFile(zip_path, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for file in file_dir:
            add_file = os.path.join(path, file)
            zf.write(add_file, file)
    yield
    os.remove(zip_path)


def test_create_arch(arc_files):
    assert zipfile.ZipFile("resources/test.zip") != None


def test_open_files(arc_files):
    with zipfile.ZipFile("resources/test.zip") as zip_file:
        print(zip_file.namelist())
        with zip_file.open("Python Testing with Pytest (Brian Okken).pdf") as pdf_file:
           reader = PdfReader(pdf_file)
           text = reader.pages[1].extract_text()
           assert "Simple, Rapid, Effective, and Scalable" in text


        with zip_file.open("example.csv", "r") as csv_file:
            reader = list(csv.reader(TextIOWrapper(csv_file)))
            print(reader[1][0])
            assert "John Doe" == reader[1][0]

        with zip_file.open("file_example_XLSX_50.xlsx") as xlsx:
            workbook = load_workbook(xlsx)
            sheet = workbook.active
            print(sheet.cell(row=1, column=1))
            assert sheet.cell(row=1, column=1).value == 0